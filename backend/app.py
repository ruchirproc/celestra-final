import io
import os
import uuid
import json
import re
import time
from pathlib import Path

from flask import Flask, request, jsonify, Response, stream_with_context
from flask_cors import CORS
from dotenv import load_dotenv
from openai import OpenAI, AzureOpenAI
import openpyxl
import httpx

load_dotenv()

app = Flask(__name__)

_ALLOWED_ORIGINS = [
    re.compile(r"http://localhost(:\d+)?$"),
    "https://celestra-final.vercel.app"
]
if _frontend_url := os.environ.get("FRONTEND_URL"):
    _ALLOWED_ORIGINS.append(_frontend_url.rstrip("/"))

CORS(app, origins=_ALLOWED_ORIGINS, supports_credentials=False)

# Auto-detect endpoint type:
#   *.openai.azure.com  → Azure OpenAI Service  → AzureOpenAI client (needs api_version)
#   *.models.ai.azure.com → Azure AI Foundry serverless → standard OpenAI client (needs /v1 suffix)
_endpoint = os.environ["AZURE_ENDPOINT"]
if "openai.azure.com" in _endpoint:
    client = AzureOpenAI(
        api_key=os.environ["AZURE_API_KEY"],
        azure_endpoint=_endpoint,
        api_version=os.environ.get("AZURE_API_VERSION", "2025-03-01-preview"),
    )
else:
    client = OpenAI(
        api_key=os.environ["AZURE_API_KEY"],
        base_url=_endpoint,   # e.g. https://xxx.eastus.models.ai.azure.com/v1
    )

MODEL = os.environ.get("AZURE_MODEL", "gpt-5.3-chat")
ASSISTANT_MODEL = os.environ.get("AZURE_ASSISTANT_MODEL", MODEL)

BASE_DIR = Path(__file__).parent.parent
AGENTS_DIR = BASE_DIR / ".claude" / "agents"
SKILLS_DIR = BASE_DIR / ".claude" / "skills"

# In-memory stores — replace with Redis/DB for production
sessions: dict = {}   # session_id → { system_prompt, messages, agent, project_id }
projects: dict = {}   # project_id → { context_json }

AGENT_SKILL_MAP = {
    "context":   ("agent-project-context.md", None),
    "targeting": ("agent-targeting.md", "Targeting"),
    "sizing":    ("agent-sizing.md", "Sizing"),
    "alignment": ("agent-alignment.md", "Alignment"),
}

_AGENT_OPENING = {
    "context":   "Hello, I am ready to begin gathering project context.",
    "sizing":    "Hello, I am ready to begin the field force sizing analysis.",
    "alignment": "Hello, I am ready to begin the territory alignment analysis.",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

_FRONTMATTER_RE = re.compile(r"^---[\s\S]*?---\s*\n?", re.MULTILINE)

# Prepended to Chat Completions system prompts (no tool access).
_API_CONTEXT = """\
You are operating as an AI assistant via API — not inside an IDE or shell.
Important constraints:
- You have NO access to a file system, terminal, or shell. When instructions say \
"save a file" or "write to disk", output the file content in your response as a \
clearly labelled code block instead.
- You have NO web search tool. When instructions say "run a web search", use your \
training knowledge and state that explicitly.
- You DO have access to any data the user pastes or uploads in the conversation.
- Always produce your complete output as text in your response.

"""

# Prepended to Responses API system prompts (Code Interpreter available).
_RESPONSES_CONTEXT = """\
You are operating as an AI assistant via the Responses API with Code Interpreter enabled.
You CAN execute Python code to read Excel files, score and tier HCPs, and write output files.
You have NO web search tool — use your training knowledge for any research needed.
When you produce an output Excel file, write it to disk so it is returned as a downloadable file.

"""


def _strip_frontmatter(text: str) -> str:
    """Remove the YAML --- block that Claude Code uses for agent/skill metadata."""
    return _FRONTMATTER_RE.sub("", text, count=1).strip()


def build_system_prompt(agent_file: str, skill_folder: str | None, responses_api: bool = False) -> str:
    agent_text = _strip_frontmatter(
        (AGENTS_DIR / agent_file).read_text(encoding="utf-8")
    )

    if skill_folder:
        skill_text = _strip_frontmatter(
            (SKILLS_DIR / skill_folder / "SKILL.md").read_text(encoding="utf-8")
        )
        agent_text = re.sub(r"<skill>[^<]+</skill>", skill_text, agent_text, count=1)

    prefix = _RESPONSES_CONTEXT if responses_api else _API_CONTEXT
    return prefix + agent_text


def _extract_text_from_response(response) -> str:
    text = ""
    for item in response.output:
        if item.type == "message":
            for part in item.content:
                if part.type == "output_text":
                    text += part.text
    return text


def parse_excel_to_text(file_bytes: bytes, max_rows: int = 500) -> str:
    """
    Convert every sheet of an Excel workbook to tab-delimited text.
    Caps at max_rows to stay within token-per-minute limits.
    The model sees enough rows to understand structure and assign scoring logic.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                all_rows.append("\t".join("" if cell is None else str(cell) for cell in row))
        if not all_rows:
            continue
        total = len(all_rows)
        display = all_rows[:max_rows]
        header = f"=== Sheet: {name} === (total rows: {total})"
        body = "\n".join(display)
        note = (
            f"\n[... {total - max_rows} additional rows omitted — "
            f"file has {total} rows total. "
            "Apply the scoring logic you derive to all rows in the full file.]"
            if total > max_rows else ""
        )
        parts.append(header + "\n" + body + note)
    return "\n\n".join(parts) if parts else "(empty workbook — no data rows found)"


def _messages_for_openai(system_prompt: str, history: list) -> list:
    """
    OpenAI expects the system prompt as the first message with role 'system',
    followed by the conversation history.
    """
    return [{"role": "system", "content": system_prompt}] + history


def _call_gpt_sync(session_id: str) -> str:
    sess = sessions[session_id]
    response = client.chat.completions.create(
        model=MODEL,
        max_completion_tokens=8192,
        messages=_messages_for_openai(sess["system_prompt"], sess["messages"]),
    )
    reply = response.choices[0].message.content or ""
    sess["messages"].append({"role": "assistant", "content": reply})
    return reply


def _save_project_context(session_id: str, reply: str) -> str:
    # Walk the reply with a bracket-depth counter to find all top-level JSON objects.
    # The old [^{}]* regex forbade nested braces and never matched real JSON.
    candidates: list[str] = []
    i = 0
    while i < len(reply):
        if reply[i] == "{":
            depth, start = 0, i
            for j in range(i, len(reply)):
                if reply[j] == "{":
                    depth += 1
                elif reply[j] == "}":
                    depth -= 1
                    if depth == 0:
                        candidates.append(reply[start : j + 1])
                        i = j
                        break
        i += 1

    # Prefer the first parseable block that contains "project_context".
    for raw in candidates:
        if '"project_context"' in raw:
            try:
                json.loads(raw)
                pid = sessions[session_id].get("project_id") or uuid.uuid4().hex
                projects[pid] = {"context_json": raw}
                sessions[session_id]["project_id"] = pid
                return pid
            except json.JSONDecodeError:
                continue

    # Fallback: first parseable JSON object.
    for raw in candidates:
        try:
            json.loads(raw)
            pid = sessions[session_id].get("project_id") or uuid.uuid4().hex
            projects[pid] = {"context_json": raw}
            sessions[session_id]["project_id"] = pid
            return pid
        except json.JSONDecodeError:
            continue

    return ""


def _format_context_for_prompt(ctx_json: str) -> str:
    """Convert the project context JSON into a readable prompt section."""
    try:
        data = json.loads(ctx_json)
    except (json.JSONDecodeError, TypeError):
        data = {}

    if not data:
        return ""

    pc = data.get("project_context", {})
    rec = data.get("claude_recommendations", {})

    lines = ["## PROJECT CONTEXT (pre-populated — use this throughout the session)"]

    drug = pc.get("drug_and_indication", {})
    if drug:
        lines.append(f"**Drug:** {drug.get('brand_name', '')} ({drug.get('generic_name', '')}) — {drug.get('indication', '')}")
        if drug.get("regulatory_status"):
            lines.append(f"**Regulatory status:** {drug['regulatory_status']}")

    ta = pc.get("therapeutic_area", {})
    if ta:
        lines.append(f"**Therapeutic area:** {ta.get('area', '')} | Rare disease: {ta.get('rare_disease', '')}")
        if ta.get("patient_population"):
            lines.append(f"**Patient population:** {ta['patient_population']}")

    comp = pc.get("competitive_landscape", {})
    if comp:
        comps = comp.get("competitors", [])
        if comps:
            lines.append(f"**Competitors:** {', '.join(comps)}")
        if comp.get("competitive_positioning"):
            lines.append(f"**Positioning:** {comp['competitive_positioning']}")

    goals = pc.get("commercial_goals", {})
    if goals:
        specs = goals.get("target_specialties", [])
        if specs:
            lines.append(f"**Target specialties:** {', '.join(specs)}")
        if goals.get("geographic_focus"):
            lines.append(f"**Geography:** {goals['geographic_focus']}")

    if rec:
        t = rec.get("for_targeting", {})
        if t:
            lines.append(f"**Recommended market type:** {t.get('market_type', '')}")
            mpo = t.get("recommended_metric_priority_order", [])
            if mpo:
                lines.append(f"**Recommended metric priority:** {', '.join(mpo)}")

    if len(lines) == 1:
        # No structured keys found — fall back to the raw JSON so the model still sees it
        return f"## PROJECT CONTEXT\n{ctx_json}"

    return "\n".join(lines)


# ── Session start ─────────────────────────────────────────────────────────────

@app.route("/api/session/start", methods=["POST"])
def start_session():
    data = request.get_json(silent=True) or {}
    agent = data.get("agent", "context")
    project_id = data.get("project_id")

    if agent not in AGENT_SKILL_MAP:
        return jsonify({"error": f"Unknown agent: {agent}"}), 400

    session_id = uuid.uuid4().hex

    # ── Targeting: Responses API + Code Interpreter ───────────────────────────
    if agent == "targeting":
        base_system_prompt = build_system_prompt("agent-targeting.md", "Targeting", responses_api=True)
        # Try in-memory store first; fall back to client-supplied JSON (survives server restarts)
        ctx = projects.get(project_id or "", {}).get("context_json") or data.get("context_json") or "{}"

        # Embed context into instructions so it persists across every turn, not just the first message
        ctx_section = _format_context_for_prompt(ctx)
        system_prompt = f"{base_system_prompt}\n\n{ctx_section}"

        opening = "I am ready to begin the HCP targeting workflow. Please upload the HCP universe Excel file to get started."
        history = [{"role": "user", "content": opening}]
        try:
            response = client.responses.create(
                model=ASSISTANT_MODEL,
                tools=[{"type": "code_interpreter", "container": {"type": "auto"}}],
                instructions=system_prompt,
                input=history,
            )
        except Exception as exc:
            return jsonify({"error": f"Failed to start targeting session: {exc}"}), 500

        first_message = _extract_text_from_response(response)
        history.append({"role": "assistant", "content": first_message})

        sessions[session_id] = {
            "agent": "targeting",
            "type": "responses",
            "system_prompt": system_prompt,
            "history": history,
            "file_ids": [],      # all uploaded file IDs — passed to container each turn
            "project_id": project_id,
        }
        return jsonify({"session_id": session_id, "first_message": first_message})

    # ── Chat Completions path (context, sizing, alignment) ───────────────────
    agent_file, skill_folder = AGENT_SKILL_MAP[agent]
    base_system_prompt = build_system_prompt(agent_file, skill_folder)

    # Inject project context for sizing and alignment (like targeting does)
    if agent in ("sizing", "alignment"):
        ctx = projects.get(project_id or "", {}).get("context_json") or data.get("context_json") or "{}"
        ctx_section = _format_context_for_prompt(ctx)
        system_prompt = f"{base_system_prompt}\n\n{ctx_section}" if ctx_section else base_system_prompt
    else:
        system_prompt = base_system_prompt

    opening = _AGENT_OPENING.get(agent, "Hello, I am ready to begin.")

    sessions[session_id] = {
        "agent": agent,
        "type": "chat",
        "system_prompt": system_prompt,
        "messages": [],
        "project_id": project_id,
    }
    sessions[session_id]["messages"].append({"role": "user", "content": opening})
    try:
        first_message = _call_gpt_sync(session_id)
    except Exception as exc:
        sessions.pop(session_id, None)
        return jsonify({"error": f"Failed to start session: {exc}"}), 500
    return jsonify({"session_id": session_id, "first_message": first_message})


# ── Message (SSE streaming) ───────────────────────────────────────────────────

@app.route("/api/session/message", methods=["POST"])
def send_message():
    if request.content_type and "multipart" in request.content_type:
        session_id = request.form.get("session_id", "")
        user_text = request.form.get("message", "")
    else:
        body = request.get_json(silent=True) or {}
        session_id = body.get("session_id", "")
        user_text = body.get("message", "")

    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid or expired session"}), 400

    sess = sessions[session_id]

    # ── Responses API path (targeting) ───────────────────────────────────────
    if sess.get("type") == "responses":
        # Upload all files — accumulates across turns so the container sees everything
        new_files = request.files.getlist("file")
        for f in new_files:
            uploaded = client.files.create(
                file=(f.filename, f.read(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                purpose="assistants",
            )
            sess["file_ids"].append(uploaded.id)

        if new_files:
            names = ", ".join(f.filename for f in new_files)
            auto_msg = (
                f"I have uploaded {len(new_files)} file(s): {names}. "
                "Please analyze them and proceed with the targeting workflow."
            )
        else:
            auto_msg = "Please continue."
        msg_text = user_text or auto_msg
        sess["history"].append({"role": "user", "content": msg_text})

        # All previously uploaded files stay available to Code Interpreter every turn
        container: dict = {"type": "auto"}
        if sess["file_ids"]:
            container["file_ids"] = sess["file_ids"]

        def generate_responses():
            full_text = ""
            output_files: list = []
            try:
                final_response = None
                last_ping = time.time()
                with client.responses.stream(
                    model=ASSISTANT_MODEL,
                    tools=[{"type": "code_interpreter", "container": container}],
                    instructions=sess["system_prompt"],
                    input=sess["history"],
                ) as stream:
                    for event in stream:
                        now = time.time()
                        if now - last_ping > 15:
                            yield ": keepalive\n\n"
                            last_ping = now
                        if event.type == "response.output_text.delta":
                            delta = getattr(event, "delta", "")
                            if delta:
                                full_text += delta
                                last_ping = time.time()
                                yield f"data: {json.dumps({'delta': delta})}\n\n"
                        elif event.type == "response.completed":
                            final_response = event.response

                # Append assistant reply to history for the next turn
                sess["history"].append({"role": "assistant", "content": full_text})

                # Collect any files written by Code Interpreter
                for item in (final_response.output if final_response else []):
                    if item.type == "code_interpreter_call":
                        for out in (item.outputs or []):
                            if out.type == "image":
                                output_files.append({
                                    "file_id": out.image.file_id,
                                    "filename": f"chart_{out.image.file_id[:8]}.png",
                                })
                    elif item.type == "message":
                        for part in (item.content or []):
                            annotations = getattr(part, "annotations", None) or []
                            for ann in annotations:
                                ann_type = getattr(ann, "type", "")

                                # Responses API uses container_file_citation (not file_path)
                                # Responses API: file_id + container_id are flat on the annotation
                                if ann_type == "container_file_citation":
                                    fid = getattr(ann, "file_id", None)
                                    cid = getattr(ann, "container_id", None)
                                    fname = getattr(ann, "filename", None) or (f"output_{fid[:8]}.xlsx" if fid else "output.xlsx")
                                    if fid and cid:
                                        output_files.append({
                                            "file_id": fid,
                                            "container_id": cid,
                                            "filename": fname,
                                        })
                                # Fallback: Assistants-style file_path annotation
                                elif ann_type == "file_path":
                                    fp = getattr(ann, "file_path", None)
                                    fid = getattr(fp, "file_id", None) if fp else None
                                    if fid:
                                        raw = getattr(ann, "text", "") or ""
                                        sandbox = re.search(r"sandbox:/[^\s)\"]+", raw)
                                        fname = sandbox.group().rstrip(")").split("/")[-1] if sandbox else f"output_{fid[:8]}.xlsx"
                                        output_files.append({"file_id": fid, "container_id": None, "filename": fname})

                yield f"data: {json.dumps({'done': True, 'is_complete': False, 'project_id': sess.get('project_id'), 'output_files': output_files})}\n\n"

            except Exception as exc:
                yield f"data: {json.dumps({'error': str(exc)})}\n\n"

        return Response(
            stream_with_context(generate_responses()),
            content_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    # ── Chat Completions path (context) ───────────────────────────────────────
    content = user_text or ""
    if "file" in request.files:
        f = request.files["file"]
        try:
            excel_text = parse_excel_to_text(f.read())
        except Exception as exc:
            return jsonify({"error": f"Could not parse uploaded file: {exc}"}), 400
        content = (content + "\n\n" if content else "") + (
            f"UPLOADED FILE: {f.filename}\n\n{excel_text}"
        )

    sess["messages"].append({"role": "user", "content": content})

    def generate():
        full_text = ""
        try:
            last_ping = time.time()
            stream = client.chat.completions.create(
                model=MODEL,
                max_completion_tokens=8192,
                messages=_messages_for_openai(sess["system_prompt"], sess["messages"]),
                stream=True,
            )
            for chunk in stream:
                now = time.time()
                if now - last_ping > 15:
                    yield ": keepalive\n\n"
                    last_ping = now
                if not chunk.choices:
                    continue
                delta = chunk.choices[0].delta.content or ""
                if delta:
                    full_text += delta
                    last_ping = time.time()
                    yield f"data: {json.dumps({'delta': delta})}\n\n"

            sess["messages"].append({"role": "assistant", "content": full_text})

            is_complete = False
            pid = sess.get("project_id")
            if sess["agent"] == "context" and "PROJECT_CONTEXT_OUTPUT:" in full_text:
                pid = _save_project_context(session_id, full_text)
                is_complete = True

            yield f"data: {json.dumps({'done': True, 'is_complete': is_complete, 'project_id': pid})}\n\n"

        except Exception as exc:
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Context form analysis ─────────────────────────────────────────────────────

@app.route("/api/context/analyze", methods=["POST"])
def analyze_context_form():
    data = request.get_json(silent=True) or {}
    form = data.get("form_data", {})
    prompt = _format_form_as_prompt(form)
    system_prompt = build_system_prompt("agent-project-context.md", None)

    session_id = uuid.uuid4().hex
    sessions[session_id] = {
        "agent": "context",
        "system_prompt": system_prompt,
        "messages": [{"role": "user", "content": prompt}],
        "project_id": None,
    }

    def generate():
        full_text = ""
        try:
            stream = client.chat.completions.create(
                model=MODEL,
                max_completion_tokens=8192,
                messages=_messages_for_openai(system_prompt, sessions[session_id]["messages"]),
                stream=True,
            )
            for chunk in stream:
                if not chunk.choices:
                    continue
                delta = chunk.choices[0].delta.content or ""
                if delta:
                    full_text += delta
                    yield f"data: {json.dumps({'delta': delta})}\n\n"

            sessions[session_id]["messages"].append({"role": "assistant", "content": full_text})
            pid = _save_project_context(session_id, full_text)

            yield f"data: {json.dumps({'done': True, 'project_id': pid})}\n\n"

        except Exception as exc:
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _format_form_as_prompt(form: dict) -> str:
    sections = []

    def section(title: str, lines: list[str]) -> None:
        filled = [l for l in lines if l]
        if filled:
            sections.append(f"{title}:\n" + "\n".join(filled))

    section("DRUG & INDICATION", [
        f"Brand name: {form['brandName']}" if form.get("brandName") else "",
        f"Generic/INN name: {form['genericName']}" if form.get("genericName") else "",
        f"Mechanism of action: {form['mechanismOfAction']}" if form.get("mechanismOfAction") else "",
        f"Primary indication: {form['indication']}" if form.get("indication") else "",
        f"Regulatory status: {form['regulatoryStatus']}" if form.get("regulatoryStatus") else "",
    ])
    section("THERAPEUTIC AREA & PATIENT POPULATION", [
        f"Therapeutic area: {form['therapeuticArea']}" if form.get("therapeuticArea") else "",
        f"Target patient population: {form['patientPopulation']}" if form.get("patientPopulation") else "",
        f"US prevalence estimate: {form['prevalenceEstimate']}" if form.get("prevalenceEstimate") else "",
        f"Rare disease: {form['rareDisease']}" if form.get("rareDisease") else "",
    ])
    competitors = form.get("competitors", [])
    section("COMPETITIVE LANDSCAPE", [
        f"Known competitors: {', '.join(competitors)}" if competitors else "",
        f"Market leader: {form['marketLeader']}" if form.get("marketLeader") else "",
        f"Competitive positioning: {form['competitivePositioning']}" if form.get("competitivePositioning") else "",
    ])
    milestones = form.get("keyMilestones", [])
    section("LAUNCH TIMELINE & MILESTONES", [
        f"Expected US launch date: {form['launchDate']}" if form.get("launchDate") else "",
        f"Commercial stage: {form['commercialStage']}" if form.get("commercialStage") else "",
        f"Key milestones: {', '.join(milestones)}" if milestones else "",
    ])
    specialties = form.get("targetSpecialties", [])
    section("COMMERCIAL GOALS & FIELD FORCE", [
        f"Target HCP specialties: {', '.join(specialties)}" if specialties else "",
        f"Geographic focus: {form['geographicFocus']}" if form.get("geographicFocus") else "",
        f"Field force model: {form['fieldForceModel']}" if form.get("fieldForceModel") else "",
        f"Initial rep count target: {form['repCountTarget']}" if form.get("repCountTarget") else "",
        f"Target deployment year: {form['deploymentYear']}" if form.get("deploymentYear") else "",
    ])

    body = "\n\n".join(sections)
    return (
        "I have gathered all project context information upfront. Here are the details:\n\n"
        + body
        + "\n\nPlease process all of this and:\n"
        "1. Provide your consulting observations for each section\n"
        "2. Derive the claude_recommendations for targeting, sizing, and alignment\n"
        "3. Output the complete project context JSON per your Step 7 instructions\n"
        "4. End with the PROJECT_CONTEXT_OUTPUT block per Step 8\n\n"
        "Do not ask clarifying questions — analyse everything as given."
    )


# ── Utility endpoints ─────────────────────────────────────────────────────────

@app.route("/api/file/<file_id>", methods=["GET"])
def download_file(file_id: str):
    container_id = request.args.get("container_id")
    filename = request.args.get("filename") or f"{file_id}.bin"

    ct_map = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls":  "application/vnd.ms-excel",
        "csv":  "text/csv",
        "png":  "image/png",
        "jpg":  "image/jpeg",
        "jpeg": "image/jpeg",
        "pdf":  "application/pdf",
    }

    def content_type_for(fname: str) -> str:
        ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
        return ct_map.get(ext, "application/octet-stream")

    try:
        if container_id:
            # Container files (cfile_*) from Responses API Code Interpreter.
            # The SDK exposes .content as a sub-resource object (not callable),
            # so we hit the REST endpoint directly.
            api_version = os.environ.get("AZURE_API_VERSION", "2025-03-01-preview")
            url = f"{_endpoint.rstrip('/')}/openai/containers/{container_id}/files/{file_id}/content"
            with httpx.Client(timeout=60) as http:
                r = http.get(url, params={"api-version": api_version},
                             headers={"api-key": os.environ["AZURE_API_KEY"]})
                r.raise_for_status()
            data = r.content
        else:
            # Standard Files API upload
            info = client.files.retrieve(file_id)
            filename = getattr(info, "filename", None) or filename
            data = client.files.content(file_id).read()

        return Response(
            data,
            content_type=content_type_for(filename),
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/project/<project_id>", methods=["GET"])
def get_project(project_id: str):
    if project_id not in projects:
        return jsonify({"error": "Not found"}), 404
    return jsonify(projects[project_id])


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "model": MODEL})


@app.errorhandler(Exception)
def handle_unhandled_exception(exc):
    # Registering a custom handler causes Flask to run after_request hooks
    # (including Flask-CORS) on the error response, so CORS headers are present.
    app.logger.exception("Unhandled exception")
    response = jsonify({"error": str(exc)})
    response.status_code = 500
    return response


if __name__ == "__main__":
    app.run(debug=True, port=5000)
